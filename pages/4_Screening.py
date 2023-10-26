##### streamlit app

import streamlit as st
import time
import joblib
import simpy

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go

import openpyxl
import tempfile

# from models.VOVO import *
import random



# Data

d ={
    'location' : 'APAC', 
    'role' : 'SWE', 
    'ac_pi_ptr' : 0.83, 
    'pi_os_ptr' : 0.33, 
    'os_oe_ptr' : 0.18, 
    'oe_oa_ptr' : 0.87, 
    'ac_pi_lat' : 4, 
    'pi_os_lat' : 4, 
    'os_oe_lat' : 4, 
    'oe_oa_lat' : 0, 
    'ppr' : 4
}

data = pd.DataFrame([d])

# goals (in OA)

goals = pd.DataFrame(
    [
        {'date': pd.to_datetime('2023/06/26'), 'quantity': 202, 'event': 'oa'}, 
        {'date': pd.to_datetime('2023/09/25'), 'quantity': 438, 'event': 'oa'}, 
        {'date': pd.to_datetime('2023/12/25'), 'quantity': 1000, 'event': 'oa'}, 
    ]
)

rng = pd.date_range(
    start = pd.to_datetime('2023/03/13'), 
    end = pd.to_datetime('2023/12/25'), freq = '7D')


d2 = data.melt(id_vars = ['location', 'role'], 
          value_vars = [
              'ac_pi_ptr', 
                'pi_os_ptr', 
                'os_oe_ptr', 
                'oe_oa_ptr', 
                'ac_pi_lat', 
                'pi_os_lat', 
                'os_oe_lat', 
                'oe_oa_lat', 
          ]
         )

v2 = d2.variable.str.split("_")

d2 = d2.assign(
    stat = v2
)

d2['x'] = d2.stat.str[0]
d2['y'] = d2.stat.str[1]
d3 = d2.groupby(['location', 'role', 'x', 'y']).apply(lambda g : g['value'].values)

d3.name = 'stats'
d3 = d3.reset_index()

st.set_page_config(
    page_title="OA Target Headcount Model",
    page_icon="🧊",
    layout="wide",
    initial_sidebar_state="expanded",
)

if 'change_dates' not in st.session_state: 
    st.session_state['change_dates'] = {}
    
if 'count' not in st.session_state:
    st.session_state['count'] = 0

# import models

# attrtition model

attrition_model = joblib.load("./models/attrition_model.joblib")

# define some re-used variables
start_date = pd.to_datetime('today').to_period('W').start_time + pd.Timedelta(6, unit='d')
end_date = pd.to_datetime('12/31/2024').to_period('W').start_time + pd.Timedelta(6, unit='d')

class Team: 
    def __init__(self, env, attrition_model, ramping_function): 
        self.Headcount = simpy.Container(env)
        self.Capacity = simpy.Container(env)
        self.attrition_model = attrition_model
        self.ramping_function = ramping_function
        
        
def attrit(env, N, record, date, team, n_weeks, init): 

    month = date.month
    if init == 1:
        pass
    else:
        team.Headcount.put(N)
        team.Capacity.put(team.ramping_function(n_weeks) * N)
    record[env.now, [n_weeks, -2]] += N
    while N > 0:
        
        n_weeks += 1 # increase week
        date += pd.Timedelta(1, unit = 'W')
        month = date.month


        l = [[n_weeks, month]]
        p = team.attrition_model.predict_proba(l)[0]

        N_new = int(N*p[0])#binom.rvs(N, p)[0]
        leaving = N - N_new
        staying = N_new
        N = N_new
                
        if leaving > 0:
            team.Headcount.get(leaving)
            team.Capacity.get(team.ramping_function(n_weeks - 1) * leaving)
            
        if staying > 0: 
            team.Capacity.get(team.ramping_function(n_weeks - 1) * staying)
            team.Capacity.put(team.ramping_function(n_weeks) * staying)
            
        yield env.timeout(1)
        
        record[env.now, [n_weeks, -2]] += N

                
def sustain(env, N, record, start_date, end_date, n_weeks, target_capacity, mode, team):   
    
    date = start_date
    month = date.month
    initial_capacity = team.ramping_function(n_weeks) * (N)
    print(initial_capacity)
    initial_headcount = N
    # N = initial_headcount

    team.Headcount.put(initial_headcount)
    record[env.now, [n_weeks-1, -1]] += N
    
    team.Capacity.put(initial_capacity)
    init = 0
#     if K > 0:   
#         env.process(
#             attrit(env, K, 0, record, date, team, 12, 1)
#         )
    
    while True: 
                
#         effective_dates = pd.date_range(start = date, end = date + pd.Timedelta(6, unit = 'D'), freq = 'D')
#         effective_changes = changes.loc[changes.change_date.isin(effective_dates)]

        team_size = team.Headcount.level
        current_capacity = team.Capacity.level 
        capacity_difference = target_capacity[env.now] - current_capacity
        
        
        
        if mode == 'capacity': 
            
            new_weeks = 3
 
            new_folks = capacity_difference//team.ramping_function(new_weeks)
            if new_folks > 0:
                record[env.now - 3, [0, -2]] += new_folks
                record[env.now - 2, [1, -2]] += new_folks
                record[env.now - 1, [2, -2]] += new_folks

        elif mode == 'headcount': 
            new_folks = target_headcount[env.now] - team_size
            
            new_weeks = 0
        else: 
            raise ValueError('Not a valid mode')
        
        if new_folks > 0: 
            init += 2
            env.process(
                attrit(env, new_folks, record, date, team, new_weeks, init)
            )
        
        n_weeks += 1
        date += pd.Timedelta(1, unit = 'W')
        month = date.month  
    

        l = [[n_weeks, month]]
        p = team.attrition_model.predict_proba(l)[0]

        N_new = int(N*p[0])#binom.rvs(N, p)[0]
        
        additional_loss = 0
#         if len(effective_changes) > 0: 
#             cohort = N_new
#             for i, row in effective_changes.iterrows():
#                 change_amount = team_size - target_headcount[env.now]
#                 layoff = np.clip(int(change_amount), a_min = 0, a_max = cohort)
#                 cohort = cohort - layoff
#                 additional_loss += layoff
            
        leaving = (N - N_new) + additional_loss
        staying = N_new - additional_loss
        N = staying
        
        if leaving > 0:
            team.Headcount.get(leaving)
            team.Capacity.get(team.ramping_function(n_weeks-1) * leaving)
            
        if staying > 0: 
            team.Capacity.get(team.ramping_function(n_weeks - 1) * staying)
            team.Capacity.put(team.ramping_function(n_weeks) * staying)


        yield env.timeout(1)
        # record the headcount of this wave
        record[env.now, [n_weeks, -1]] += N
        
        
def run_simulation(N, attrition_model, ramping_function, target_capacity, start_date, end_date, mode):
    n_weeks = 12

    sim_len = (end_date - start_date).days // 7    
    record = np.zeros((sim_len, n_weeks + sim_len + 2))

    env = simpy.Environment()
    team = Team(env, attrition_model, ramping_function)
    
    env.process(
        sustain(env, N, record, start_date, end_date, n_weeks, target_capacity = target_capacity,
                team = team, mode = mode)
    )
    env.run(sim_len)
    
    return record

ac_oa_ptr = data[data.columns[data.columns.str.endswith('ptr')]].prod(axis = 1)
ac_oa_lat = data[data.columns[data.columns.str.endswith('lat')]].sum(axis = 1)
ac_ppr = (data['ppr'] / ac_oa_ptr).values[0] / 4.6

# temporaty! define ramping function
def ramping_function(week, ppr_list): 
    for w, tickets in enumerate(ppr_list): 
        if week == w: 
            return tickets
    return ppr_list[-1]

def location_roles(location, actuals = data): 
    roles = actuals.loc[actuals.location == location].role.unique()
    return tuple(roles)

def get_initials(location, role, actuals): 
    act = actuals.loc[
        (actuals.location == location) & (actuals.role.isin(role))
    ]
    
    return act

def increment_count():
    st.session_state.count += 1
    return 
    
def render_elem(): 
    s = ''
    for row in range(st.session_state.count): 
        s += f"""
col1, col2, col3 = st.columns(3)
with col1: 
    st.markdown("Target {row + 2}")
with col2: 
    change_date{row+1} = st.date_input('Target Date', 
                                value=change_date{row} + pd.Timedelta(13, unit = 'W'),
                                min_value=change_date{row} + pd.Timedelta(13, unit = 'W'), 
                                max_value=end_date, key = 'change_date_{row+1}'
    )
with col3: 
    change_amount{row+1} = st.number_input('OA Goal', min_value=0.0, max_value=10000.0, value=0.0, key = 'change_amount_{row+1}', 
                                    label_visibility="visible")
        """
    return s

def record_change_dates(): 
    for row in range(st.session_state.count): 
        st.session_state.change_dates[row + 2] = {'date': eval(f'change_date{row+1}'), 'quantity': eval(f'change_amount{row+1}'), 'event': 'oa'}
    return 

def reset_change_dates(): 
    st.session_state.count = 0
    st.session_state.change_dates = {}
    # st.session_state.change_dates = st.session_state.change_dates[1]
    
def summarize_data(l): 
    pass
#     wb = openpyxl.load_workbook("./models/VOVO_output_template.xlsx")
#     ws = wb['NA Transition Plan']
    
#     l['Month'] = pd.to_datetime(l.date) +  pd.offsets.MonthBegin(-1)
#     l = l .loc[(l.Month >= pd.to_datetime('2023/06/01')) & (l.Month <= pd.to_datetime('2024/04/01'))]
#     li = l.groupby(['Role', 'Month'], as_index = False)['optimal_starting'].sum()
    
#     li = li.loc[(li.Month >= pd.to_datetime('2023/06/01')) & (li.Month <= pd.to_datetime('2024/04/01'))]
#     li_pivot = li.pivot(index = 'Role', columns = 'Month', values = 'optimal_starting')
    
#     l['mo_week'] = l.groupby(['Role', 'Month']).rank()['date']
#     lj = l.loc[l.mo_week == 1.]
#     lj['prev_mo_remote'] = lj.groupby('Role').remote_headcount.shift(1)
#     lj = lj.assign(
#         remote_depreciation = -1*(lj['prev_mo_remote'] - lj['remote_headcount'])
#     )
#     remote_hc_ramp_down = lj.groupby('Month')['remote_depreciation'].sum().values
    
#     starters = lj.loc[lj.date == lj.date.min()]
#     enders = lj.loc[lj.date == lj.date.max()]
    
#     for i,row in enumerate(li_pivot.iterrows()): 
#         ws[f'A{i+5}'] = row[0]
#         ws[f'B{i+5}'] = starters.loc[starters.Role == row[0]]['Total Headcount'].values[0]
#         ws[f'C{i+5}'] = starters.loc[starters.Role == row[0]]['vovo_headcount'].values[0]
#         ws[f'D{i+5}'] = starters.loc[starters.Role == row[0]]['remote_headcount'].values[0]
        
        
    
#     ws[f'P{i+5}'] = enders.loc[enders.Role == row[0]]['vovo_headcount'].values[0]
#     # remote_hc_ramp_down = lj.loc[lj.Role == row[0]]['remote_depreciation'].values
#     for col, vovo_val, remo_val in zip('EFGHIJKLMNO', row[1].values, remote_hc_ramp_down): 
#         ws[col+f'{i+5}'] = vovo_val
#         ws[col+'13'] = remo_val
        
#     wb.save('./vovo_output.xlsx')
        
#     return 
    

# @st.cache_resource
def Simulation(target_capacity, 
               ramping_function = ramping_function,
               attrition_model = attrition_model):
    
    start_date = target_capacity.index.min()
    end_date = target_capacity.index.max()
    target_capacity = target_capacity.values
    N = np.ceil(target_capacity[0] / ramping_function(5))
    mode = 'capacity'
    
    record = run_simulation(N, attrition_model, ramping_function, target_capacity, start_date, end_date, mode)
    rf = np.vectorize(ramping_function)
    period = pd.date_range(start_date, end_date, freq = 'W', inclusive = 'left')
    df = pd.DataFrame(record)
    df = df.assign(
        global_capacity = np.dot(df.iloc[:, :-2].values, rf(df.iloc[:, :-2].columns).T),
        global_headcount = (df.iloc[:, -2] + df.iloc[:, -1]), 
        vovo_headcount = df.iloc[:, -2], 
        remote_headcount = df.iloc[:, -1], 
        date = period
    )
    
    r, c = record.shape
    df['weekly_attrition'] = df[0]

    LL_agg = df.copy()

    LL_agg['month'] = LL_agg.date.dt.month
    LL_agg['Year'] = LL_agg.date.dt.year
    LL_agg['Quarter'] = np.nan
    LL_agg.loc[LL_agg.month == 1, 'Quarter'] = 1
    LL_agg.loc[LL_agg.month == 4, 'Quarter'] = 2
    LL_agg.loc[LL_agg.month == 7, 'Quarter'] = 3
    LL_agg.loc[LL_agg.month == 10, 'Quarter'] = 4

    LL_agg = LL_agg.fillna(method = 'ffill')
    LL_agg['last_q'] = LL_agg['Quarter'].shift(1)

    LL_agg['cum_tix'] = LL_agg.groupby(['Quarter', 'Year']).global_capacity.cumsum()
    LL_agg_loc = LL_agg.copy()

    LL_agg_loc['total_team_size'] = LL_agg_loc.loc[:, 0:c-1].sum(axis = 1)
    LL_agg_loc['5+'] = LL_agg_loc.loc[:, 4:c-1].sum(axis = 1)
    LL_agg_loc['date'] = pd.to_datetime(LL_agg_loc['date']).dt.date

    return LL_agg, LL_agg_loc
    # return df, initial_headcount

st.markdown("# Screening Headcount Model")
st.sidebar.markdown("""
# Control Panel
Choose the location and role. Then choose pick OA goals that the simulation will target
""")
location = st.sidebar.selectbox(
    'Location',
    ('APAC', )
)

options = location_roles(location)
role = st.sidebar.multiselect(
    'Role',
    options, 
    default = options[0]
    
)

st.sidebar.markdown(f"""
### Targets
Introduce OA goals over time 

""")

with st.sidebar.expander("Targets"): 
    col1, col2, col3 = st.columns(3)

    with col1: 
        st.markdown("Target 1")

    with col2: 
        change_date0 = st.date_input('Target Date', 
                                    value=pd.to_datetime('09/25/2023'),
                                    min_value=start_date + pd.Timedelta(3, unit = 'W'), 
                                    max_value=end_date
        )
    with col3: 
        change_amount0 = st.number_input('OA Goal', min_value=0.0, max_value=10000., value=200.0,
                                        label_visibility="visible")
    exec(render_elem())
    st.session_state.change_dates[1] = {'date': change_date0, 'quantity': change_amount0, 'event': 'oa'}
    record_change_dates()
    add_intervention = st.button('Add Target', use_container_width=True, on_click = increment_count)
    reset_interventions = st.button('Reset', use_container_width=True, on_click = reset_change_dates)
    
goals = pd.DataFrame(st.session_state['change_dates'].values())

other_goals = []
def calc_goals(goal, deadline, event): 
    
    n = d3[d3['y'] == event]
    val, l = n['stats'].values[0]
    goal /= val
    deadline -= pd.to_timedelta(l, unit = 'W')
    
    other_goals.append({'date': deadline, 'quantity': np.ceil(goal), 'event': n['x'].values[0]})
    
    try: 
        calc_goals(goal, deadline, n['x'].values[0])
    except IndexError: 
        pass

_ = goals.apply(lambda row : calc_goals(row['quantity'], row['date'], row['event']), axis = 1)
dl = goals.append(other_goals).reset_index(drop = True)

def realize(row, L): 
    dr = pd.date_range(row['date'] - pd.Timedelta(13, unit = 'W'), row['date'], freq = 'W')
    e = row['quantity'] / 13
    cl = row['event']
    
    df = pd.DataFrame()
    df['date'] = dr
    df['quantity'] = e
    df['event'] = cl
    
    L.append(df)
    return

L = []
_ = dl.apply(realize, L = L, axis = 1)

dll = pd.concat(L).groupby(['date', 'event'], as_index = False).sum()
dll2 = dll.pivot('date', 'event', 'quantity')
dll2 = dll2.fillna(method = 'ffill')

act = get_initials(location, role, actuals = data)

st.markdown("""
<style>
div[data-testid="metric-container"] {
   background-color: rgba(28, 131, 225, 0.1);
   border: 1px solid rgba(28, 131, 225, 0.1);
   padding: 5% 5% 5% 10%;
   border-radius: 5px;
   color: rgb(30, 103, 119);
   overflow-wrap: break-word;
}

/* breakline for metric text         */
div[data-testid="metric-container"] > label[data-testid="stMetricLabel"] > div {
   overflow-wrap: break-word;
   white-space: break-spaces;
   color: red;
}
</style>
"""
, unsafe_allow_html=True)
# col1, col2, col3, col4 = st.columns(4)
# with col1: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
# with col2: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
# with col3: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
# with col4: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
    
# col1, col2, col3, col4 = st.columns(4)
# with col1: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
# with col2: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
# with col3: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")
# with col4: 
#     st.metric(label="This is a very very very very very long sentence", value="70 °F")

target_capacity_ac = dll2['ac']

ac_oa_ptr = data[data.columns[data.columns.str.endswith('ptr')]].prod(axis = 1)
ac_oa_lat = data[data.columns[data.columns.str.endswith('lat')]].sum(axis = 1)
ac_ppr = (data['ppr'] / ac_oa_ptr).values[0] / 4.6
ppr_list = [n * ac_ppr for n in [0.25, 0.5, 0.75, 1]]
ramping_function_rec = lambda n : ramping_function(n, ppr_list = ppr_list)

LL_agg, LL_agg_loc = Simulation(target_capacity_ac, ramping_function = ramping_function_rec)

target_capacity_sched = dll2[['pi', 'os']].assign(
    pi = dll2.pi * 1.6, 
    os = dll2.os * 10.5
).sum(axis = 1)

target_capacity_sched = target_capacity_sched[target_capacity_sched > 0]


ppr_list = [13, 26, 39, 65]
ramping_function_sched = lambda n : ramping_function(n, ppr_list = ppr_list)
LL_agg_sched, LL_agg_loc_sched = Simulation(target_capacity_sched, ramping_function = ramping_function_sched)

# group_df = DF.groupby('date', as_index = False)['remote_headcount'].sum()

# try: 
    
#     last_day = group_df.loc[group_df.remote_headcount == 0].head(1)
#     st.markdown(f"""
#     ### {location} {workstream}: {role}
#     #### Currently, there are {int(hc)} employees outside of VOVO areas and {int(vovo_hc)} VOVO-aligned. Your interventions led to {int(last_day['remote_headcount'].values[0])} employees in the remote workforce by {last_day['date'].dt.date.values[0]}. Making sure the {mode.lower()} did not drop below baseline meant adding more new workers to the VOVO team than otherwise would have been added. Click on `Log` for more details. 
#     """)
# except IndexError: 
#     last_day = group_df.loc[group_df.remote_headcount > 0].tail(1)
#     st.markdown(f"""
#     ### {location} {workstream}: {role}
#     #### Currently, there are {int(hc)} employees outside of VOVO areas and {int(vovo_hc)} VOVO-aligned. Your interventions led to {int(last_day['remote_headcount'].values[0])} employees still in the remote workforce by {last_day['date'].dt.date.values[0]}. An optimal hiring schedule was computed such that this team's {mode.lower()} did not drop below baseline. Click on `Log` for more details. 
#     """)
with st.expander('Log') : 
    dl
    
#     f = DF[['date', 0, 'Total Capacity', 'Total Headcount', 'vovo_headcount', 'remote_headcount', 'Role']].rename(
#         columns = {0 : 'optimal_starting'}
#     )
#     f
#     st.download_button('Download this data', data = f.to_csv().encode('utf-8'), use_container_width=True)
    
#     with tempfile.TemporaryDirectory() as tmpdirname:
#         summarize_data(f)
#         print('created temporary directory', tmpdirname)

#         # directory and contents have been removed

#         with open('./vovo_output.xlm', "rb") as k:
#             st.download_button('Download the Google-approved summary of this data', data = k, use_container_width=True, 
#                                mime = "application/vnd.google-apps.spreadsheet"
#                               )

col1, col2 = st.columns(2)

with col1:
    st.markdown("""
    ### Deadlines
    """)
    
    fig = px.bar(dl.sort_values(['event', 'date'], ascending = False), x = 'date', y = 'quantity', color = 'event')
    st.plotly_chart(fig, use_container_width=True)
with col2: 
    st.markdown(f"""
    ### Required Weekly Yield 
    """)
    
    fig = px.line(dll.sort_values(['event', 'date'], ascending = False), y = 'quantity', x = 'date', color = 'event', line_shape = 'hv')
    st.plotly_chart(fig, use_container_width=True)

st.markdown("""
<h3 style='text-align: center;'>Estimated Required Headcount Over Time</h3>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1: 
    
    st.markdown("""
    #### Recruiters/Sourcers
    """)

    tab1, tab2 = st.tabs(['Raw Headcount', 'Broken Down by Experience Level'])
    
    with tab1: 
        fig = px.bar(LL_agg_loc, x = 'date', y = 'total_team_size')
        fig.update_layout(
            xaxis_title="Date", yaxis_title="Estimated Required Headcount"
        )
        st.plotly_chart(fig, use_container_width=True)
    
    with tab2: 
#         st.markdown("""
#         ### Estimated Headcount Broken Down by Productivity
#         """)
        fig = px.bar(LL_agg_loc, x = 'date', y = [0, 1, 2, 3, '5+'])
        fig.update_layout(
            xaxis_title="Date", yaxis_title="Headcount"
        )
        st.plotly_chart(fig, use_container_width=True)
with col2: 
    st.markdown("""
    #### Schedulers
    """)
    tab1, tab2 = st.tabs(['Raw Headcount', 'Broken Down by Experience Level'])
    
    with tab1: 
        fig = px.bar(LL_agg_loc_sched, x = 'date', y = 'total_team_size')
        fig.update_layout(
            xaxis_title="Date", yaxis_title="Estimated Required Headcount"
        )
        st.plotly_chart(fig, use_container_width=True)
    
    with tab2: 
#         st.markdown("""
#         ### Estimated Headcount Broken Down by Productivity
#         """)
        fig = px.bar(LL_agg_loc_sched, x = 'date', y = [0, 1, 2, 3, '5+'])
        fig.update_layout(
            xaxis_title="Date", yaxis_title="Headcount"
        )
        st.plotly_chart(fig, use_container_width=True)

LL_agg_loc = LL_agg_loc[['date', 'global_capacity', 'Year', 'Quarter']].assign(
    oas_achieved = (LL_agg_loc.global_capacity * ac_oa_ptr.values[0]).shift(ac_oa_lat.values[0])
)

dll2 = dll2.reset_index()
dll2['month'] = dll2.date.dt.month
dll2['Year'] = dll2.date.dt.year
dll2['Quarter'] = np.nan
dll2.loc[dll2.month == 1, 'Quarter'] = 1
dll2.loc[dll2.month == 4, 'Quarter'] = 2
dll2.loc[dll2.month == 7, 'Quarter'] = 3
dll2.loc[dll2.month == 10, 'Quarter'] = 4
dll2['Quarter'] = dll2['Quarter'].fillna(method = 'ffill')

dll2['Cummulative Offers Excepted'] = dll2.groupby(['Year', 'Quarter'])['oa'].cumsum()
LL_agg_loc['Estimated Cummulative Offers Excepted'] = LL_agg_loc.groupby(['Year', 'Quarter'])['oas_achieved'].cumsum()
fig = px.line(LL_agg_loc, x = 'date', y = 'Estimated Cummulative Offers Excepted', line_shape = 'hv')
fig.add_trace(
    go.Scatter(x = dll2.date, y = dll2['Cummulative Offers Excepted'], mode='lines', name = 'Target Weekly OAs', line=dict(
        shape='hv'
    ), 
              line_color='black')
)
# fig.update_yaxes(range=[0, 2.0 * initial_headcount])
fig.update_layout(
    xaxis_title="Date", yaxis_title="Headcount"
)
st.plotly_chart(fig, use_container_width=True)